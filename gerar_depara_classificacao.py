"""Gera o de-para de classificacao de venda (Proposta -> 14 colunas manuais) a
partir da planilha oficial de fechamento (Vendas Consolidadas.xlsm).

As colunas de classificacao (Origem, Canal, Midia, E Lead?, Diretoria/House-
Parcerias, on/off, Pago ou Organico, Qtd Corretores, Reciclagem?/Qtd, Ativacao?/
Qtd, Perdeu Roleta?/Qtd, Share, Modulo) NAO vem do CVDW: sao preenchidas a mao no
fechamento. Este script CONSOLIDA as abas espalhadas do .xlsm numa UNICA tabela
canonica, chaveada por Proposta (= idreserva), e escreve de_para_classificacao.xlsx.

Precedencia quando a mesma proposta aparece em varias abas:
  linha mais COMPLETA vence (mais colunas preenchidas); empate -> aba mais recente.

Uso:
  python gerar_depara_classificacao.py [--xlsm <caminho>] [--saida <arquivo.xlsx>]
  (sem --xlsm usa VENDAS_CONSOLIDADAS_XLSM do .env)
"""
from __future__ import annotations

import argparse
import os
import re
import unicodedata
from pathlib import Path

import openpyxl

from config.settings import RAIZ, carregar_config_pg

# Abas conhecidas que carregam classificacao, com prioridade de recencia (maior = vence no empate).
PRIORIDADE_ABAS = {
    "Vendas Novas": 100,
    "Consolidado": 95,
    "Detalhes1": 90,
    "Detalhes2": 89,
    "Vendas Consolidadas": 80,
    "Vendas - 2024 (CV)": 70,
    "Vendas - 2023": 60,
}
PRIORIDADE_PADRAO = 10  # qualquer outra aba auto-detectada

# header normalizado -> nome canonico de saida
TARGETS = {
    "origem": "Origem", "origem2": "Origem2", "canal": "Canal", "midia": "Mídia",
    "e lead?": "É Lead?", "e lead": "É Lead?",
    "diretoria / house-parcerias": "Diretoria / House-Parcerias",
    "diretoria/house-parcerias": "Diretoria / House-Parcerias",
    "on/off": "on/off", "pago ou organico": "Pago ou Orgânico",
    "corretores": "Corretores", "qtd corretores": "Qtd Corretores",
    "regiao": "Região", "share": "Share", "modulo": "Módulo",
    "reciclagem?": "Reciclagem?", "reciclagem": "Reciclagem?", "qtd reciclagem": "Qtd Reciclagem",
    "ativacao?": "Ativação?", "ativacao": "Ativação?", "qtd ativacao": "Qtd Ativação",
    "perdeu roleta?": "Perdeu Roleta?", "perdeu roleta": "Perdeu Roleta?",
    "qtd perdeu roleta": "Qtd perdeu roleta",
}
OUT_COLS = [
    "Proposta", "Origem", "Origem2", "Canal", "Mídia", "É Lead?",
    "Diretoria / House-Parcerias", "on/off", "Pago ou Orgânico", "Corretores",
    "Qtd Corretores", "Região", "Share", "Módulo", "Reciclagem?", "Qtd Reciclagem",
    "Ativação?", "Qtd Ativação", "Perdeu Roleta?", "Qtd perdeu roleta",
]
CLASSIF_COLS = OUT_COLS[1:]


def norm(s) -> str:
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


def como_int(v):
    if v is None:
        return None
    m = re.match(r"\s*(\d+)", str(v))
    return int(m.group(1)) if m else None


def mapear_aba(ws):
    """Acha (linha_header, col_chave, {canonico: idx}) ou None se a aba nao serve."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        if not row:
            continue
        nmap = {norm(c): j for j, c in enumerate(row) if c is not None}
        chave = nmap.get("proposta", nmap.get("reserva"))
        alvos = {TARGETS[n]: j for n, j in nmap.items() if n in TARGETS}
        if chave is not None and len(alvos) >= 2:
            return i, chave, alvos
    return None


def coletar(xlsm: Path) -> dict[int, dict]:
    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    melhor: dict[int, tuple] = {}  # proposta -> (completude, prioridade, {col: val})
    for ws in wb.worksheets:
        cfg = mapear_aba(ws)
        if not cfg:
            continue
        hidx, kcol, alvos = cfg
        prio = PRIORIDADE_ABAS.get(ws.title, PRIORIDADE_PADRAO)
        n = 0
        for row in ws.iter_rows(min_row=hidx + 2, values_only=True):
            if not row or kcol >= len(row):
                continue
            prop = como_int(row[kcol])
            if not prop:
                continue
            vals = {}
            for canon, j in alvos.items():
                if j < len(row) and row[j] not in (None, ""):
                    vals[canon] = row[j]
            if not vals:
                continue
            compl = len(vals)
            atual = melhor.get(prop)
            if atual is None or (compl, prio) > (atual[0], atual[1]):
                melhor[prop] = (compl, prio, vals)
            n += 1
        if n:
            print(f"  {ws.title[:32]:34} <- {n:5d} linhas (prio {prio})")
    return {p: v[2] for p, v in melhor.items()}


def cobertura(propostas: set[int]) -> None:
    try:
        from cvdw import db
        with db.conectar(carregar_config_pg()) as conn, conn.cursor() as cur:
            cur.execute("SELECT idreserva::bigint FROM bronze.reservas")
            res = {int(r[0]) for r in cur.fetchall()}
    except Exception as exc:  # noqa: BLE001
        print(f"  (cobertura: sem banco — {str(exc).splitlines()[0]})")
        return
    classif = res & propostas
    print(f"\nCobertura vs bronze.reservas ({len(res)} reservas):")
    print(f"  classificadas: {len(classif)}  ({100*len(classif)//max(len(res),1)}%)")
    print(f"  em branco    : {len(res - propostas)}")
    print(f"  no de-para mas fora do bronze: {len(propostas - res)}")


def main() -> int:
    ap = argparse.ArgumentParser(description="Gera de_para_classificacao.xlsx da Vendas Consolidadas.")
    ap.add_argument("--xlsm", help="caminho do .xlsm (default: VENDAS_CONSOLIDADAS_XLSM do .env)")
    ap.add_argument("--saida", help="arquivo de saida (default: DEPARA_DIR/de_para_classificacao.xlsx do .env)")
    args = ap.parse_args()

    dir_saida = os.getenv("DEPARA_DIR")
    saida_padrao = (Path(dir_saida) if dir_saida else RAIZ / "powerbi") / "de_para_classificacao.xlsx"
    args.saida = args.saida or str(saida_padrao)

    xlsm = Path(args.xlsm or os.getenv("VENDAS_CONSOLIDADAS_XLSM", ""))
    if not xlsm.exists():
        print(f"ERRO: xlsm nao encontrado: {xlsm}")
        return 2

    print(f"Lendo {xlsm.name} ...")
    dados = coletar(xlsm)
    print(f"\nTotal consolidado: {len(dados)} propostas distintas.")

    # colunas com dados (diagnostico de cobertura por coluna)
    print("Preenchimento por coluna:")
    for c in CLASSIF_COLS:
        n = sum(1 for v in dados.values() if v.get(c) not in (None, ""))
        if n:
            print(f"  {c:32} {n:5d}")

    saida = Path(args.saida)
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "de_para_classificacao"
    ws.append(OUT_COLS)
    for prop in sorted(dados):
        v = dados[prop]
        ws.append([prop] + [v.get(c) for c in CLASSIF_COLS])
    wb.save(saida)
    print(f"\nGravado: {saida}")

    cobertura(set(dados))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
