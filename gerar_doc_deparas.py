"""Gera um Excel de documentação dos de-paras a partir do config/deparas.yml.
Escreve <DEPARA_DIR>/de_paras_documentacao.xlsx (README da pasta de de-paras).

  python gerar_doc_deparas.py
"""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config.settings import RAIZ

MANIFESTO = RAIZ / "config" / "deparas.yml"

# grupo (ordem/cor) + momento + pra que serve — a parte editorial (não está no yml).
USO = {
    "classificacao":            ("Ativo",   "Power BI (merge por Proposta) — agora", "origem, canal, mídia, é lead, house-parcerias, on/off, Região 3-way, Share, reciclagem, ativação, roleta"),
    "empreendimento_regional":  ("Ativo",   "silver → gold.fato_reservas",           "regional RPO/URA/SCA/MTO (do empreendimento)"),
    "empreendimento_produtos":  ("Ativo",   "silver",                                 "nome de empreendimento conformado ('Fiusa 016')"),
    "situacao_esteira":         ("Ativo",   "silver → gold",                          "situação tratada + ordem do funil (esteira)"),
    "ordem_etapa":              ("Ativo",   "silver → gold",                          "ordem do funil de situação"),
    "corretor_fora_ranking":    ("Ativo",   "Power BI (filtro no ranking)",           "exclui corretores de coordenação do ranking"),
    "responsavel_imobiliaria":  ("Ativo",   "silver",                                 "responsável → imobiliária"),
    "empreendimentos":          ("A usar",  "dimensão empreendimento",                "regional completa, viabilidade, IVV — carregar fecha as 122 reservas sem regional"),
    "metas":                    ("A usar",  "dashboard meta × realizado",             "metas de QTD/VGV por empreendimento"),
    "canal_midia":              ("A usar",  "modelagem de LEADS",                     "canal 2.0 / mídia conformada (funil de leads)"),
    "ativo_receptivo":          ("A usar",  "modelagem de LEADS",                     "canal → Ativo/Receptivo/Diretoria"),
    "qualificacao_lead":        ("A usar",  "modelagem de LEADS",                     "situação do lead → qualificado (MQL)"),
    "headcount":                ("A usar",  "CAC / produtividade",                    "headcount por equipe/mês"),
    "estrutura_precos":         ("A usar",  "preço",                                  "estrutura/preço base por empreendimento e unidade"),
    "equipe_corretor":          ("A usar",  "corretores",                             "corretor → categoria/equipe"),
    "etapa_precadastro":        ("Pendente","modelagem de PRÉ-CADASTRO",              "etapa WKF → etapa BI do funil de crédito (falta xlsx)"),
    "profissoes":               ("Pendente","perfil de cliente",                      "profissão → micro/macro (falta xlsx)"),
    "feriados":                 ("Pendente","tempos médios / SLA",                    "calendário de feriados em dias úteis (falta xlsx)"),
    "gerentes":                 ("Legado",  "— (ranking_gerentes removido; House vem da classificação oficial)", "gerente → share/house/regional (sem consumidor no gold; candidato a descarte)"),
    "imobiliaria_house":        ("Legado",  "— (saiu da fato no Phase 2)",            "imobiliária → house/regional (candidato a descarte)"),
}
ORDEM = {"Ativo": 0, "A usar": 1, "Pendente": 2, "Legado": 3}
COR = {"Ativo": "D8EAD3", "A usar": "FCE9C8", "Pendente": "F8D0CE", "Legado": "E6E6E6"}


def como_atualizar(tipo: str) -> str:
    return {
        "arquivo":  "substituir o .xlsx pela versão nova da fonte (SharePoint)",
        "silver":   "rodar montar_estrutura_depara.py (re-exporta do silver)",
        "gerado":   "rodar gerar_depara_classificacao.py + montar_estrutura_depara.py",
        "pendente": "falta a fonte (xlsx do SharePoint / carga)",
    }.get(tipo, "")


def main() -> int:
    cfg = yaml.safe_load(MANIFESTO.read_text(encoding="utf-8"))
    linhas = []
    for d in cfg["deparas"]:
        nome = d["nome"]
        status, momento, praque = USO.get(nome, ("?", "", d.get("descricao", "")))
        fonte = d.get("fonte_rel") or ("materializado do silver" if d["tipo"] == "silver"
                                       else "pendente" if d["tipo"] == "pendente" else "—")
        linhas.append((ORDEM.get(status, 9), status, f"depara_{nome}", momento, praque,
                       d.get("silver") or "—", fonte, como_atualizar(d["tipo"])))
    linhas.sort(key=lambda x: (x[0], x[2]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "de-paras"
    cols = ["Status", "De-para", "Momento / camada", "Pra quê (colunas/uso)",
            "Tabela silver", "Arquivo (fonte)", "Como atualizar"]
    ws.append(cols)

    hdr_fill = PatternFill("solid", fgColor="003254")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="C9D2DA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
        c.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 22

    for _, status, nome, momento, praque, silver, fonte, atualizar in linhas:
        ws.append([status, nome, momento, praque, silver, fonte, atualizar])
        fill = PatternFill("solid", fgColor=COR.get(status, "FFFFFF"))
        for c in ws[ws.max_row]:
            c.fill, c.border = fill, border
            c.alignment = Alignment(vertical="top", wrap_text=True)

    larguras = [11, 26, 30, 52, 26, 40, 46]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}{ws.max_row}"

    destino = Path(os.getenv("DEPARA_DIR", str(RAIZ / "powerbi"))) / "de_paras_documentacao.xlsx"
    wb.save(destino)
    print(f"Gravado: {destino}  ({len(linhas)} de-paras)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
