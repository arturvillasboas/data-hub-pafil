"""Popula os seeds silver.dpara_* a partir do JSON base64+DEFLATE embutido no Power
Query legado (_bi_ref/M_Empreendimentos.md) e, opcionalmente, de planilhas .xlsx.

  python popular_seeds.py [--gerentes depara_gerentes.xlsx] [--xlsm "Vendas Consolidadas.xlsm"]
"""
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import zlib
from pathlib import Path

from config.settings import RAIZ, carregar_config_pg
from cvdw import db
from cvdw.log import configurar_logging, get_logger

log = get_logger("popular_seeds")

MD_LEGADO = RAIZ.parent / "_bi_ref" / "M_Empreendimentos.md"

# expr no markdown -> como carregar no seed.
#   json_cols : ordem das colunas em cada linha decodificada
#   tabela    : tabela alvo em silver
#   linha     : função (dict json_col->valor) -> dict (col_seed -> valor) p/ INSERT
MAPEAMENTOS = [
    dict(
        expr="dpara_ativo_receptivo",
        tabela="dpara_ativo_receptivo",
        json_cols=["canal", "lead_ou_prospect"],
        linha=lambda r: {"canal": r["canal"], "lead_ou_prospect": r["lead_ou_prospect"]},
    ),
    dict(
        expr="dpara_qualificacao_lead",
        tabela="dpara_qualificacao_lead",
        json_cols=["situacao", "qualificado"],
        linha=lambda r: {"situacao": r["situacao"], "qualificado": r["qualificado"]},
    ),
    dict(
        expr="dpara_ordem_etapa",
        tabela="dpara_ordem_etapa",
        json_cols=["etapa", "ordem"],
        linha=lambda r: {"etapa": r["etapa"],
                         "ordem": int(float(r["ordem"])) if r["ordem"] not in (None, "") else None},
    ),
    # DP-01 (gerente): removido daqui em 21/jul/2026 — o JSON legado só tinha 12/22
    # gerentes; a seed dpara_gerente_contexto vem exclusivamente do xlsx (carregar_gerentes).
    # Não há mais fallback JSON p/ esta seed.
    dict(
        # DP-02: canal/mídia (versão out24). vigência fixa.
        expr="dpara_canal_midia_out24",
        tabela="dpara_canal_midia",
        json_cols=["concat", "canal", "midia"],
        linha=lambda r: {"concat": r["concat"], "canal": r["canal"], "midia": r["midia"],
                         "vigencia_de": "2024-10-01"},
    ),
]


def _secoes(md: str) -> dict[str, str]:
    """Quebra o markdown em {nome_expr: texto} pelos cabeçalhos '### EXPR/TABELA nome'."""
    secoes: dict[str, str] = {}
    atual = None
    buffer: list[str] = []
    for linha in md.splitlines():
        m = re.match(r"^###\s+(?:EXPR|TABELA)\s+(.+?)\s*$", linha)
        if m:
            if atual:
                secoes[atual] = "\n".join(buffer)
            atual = m.group(1)
            buffer = []
        elif atual:
            buffer.append(linha)
    if atual:
        secoes[atual] = "\n".join(buffer)
    return secoes


def _decodificar(texto_secao: str) -> list[list]:
    """Extrai o 1º Binary.FromText e devolve as linhas JSON (lista de listas)."""
    m = re.search(r'Binary\.FromText\("([^"]+)"', texto_secao)
    if not m:
        raise ValueError("nenhum Binary.FromText na seção")
    raw = zlib.decompress(base64.b64decode(m.group(1)), -15)  # -15 = raw DEFLATE
    return json.loads(raw)


def carregar(conn, mapa: dict) -> int:
    secoes = carregar.secoes  # type: ignore[attr-defined]
    if mapa["expr"] not in secoes:
        log.warning("  expr %s não encontrada no markdown — pulando.", mapa["expr"])
        return 0
    linhas_json = _decodificar(secoes[mapa["expr"]])

    registros: list[dict] = []
    for valores in linhas_json:
        r = dict(zip(mapa["json_cols"], valores))
        registros.append(mapa["linha"](r))
    if not registros:
        return 0

    colunas = list(registros[0].keys())
    tabela = mapa["tabela"]
    from psycopg import sql

    placeholders = sql.SQL(", ").join(sql.Placeholder() for _ in colunas)
    consulta = sql.SQL(
        "INSERT INTO silver.{tab} ({cols}) VALUES ({ph}) ON CONFLICT DO NOTHING"
    ).format(
        tab=sql.Identifier(tabela),
        cols=sql.SQL(", ").join(map(sql.Identifier, colunas)),
        ph=placeholders,
    )
    with conn.cursor() as cur:
        cur.execute(sql.SQL("TRUNCATE silver.{}").format(sql.Identifier(tabela)))
        inseridos = 0
        for reg in registros:
            cur.execute(consulta, [reg[c] for c in colunas])
            inseridos += cur.rowcount
    log.info("  silver.%-22s <- %3d linhas (de %d no JSON)", tabela, inseridos, len(linhas_json))
    return inseridos


def _tabela_openpyxl(ws, nome_tabela: str) -> list[dict]:
    """Lê uma Excel Table nomeada (ws.tables) e devolve as linhas como dicts
    {cabeçalho: valor}, pulando linhas totalmente vazias."""
    cells = ws[ws.tables[nome_tabela].ref]
    header = [c.value for c in cells[0]]
    linhas = []
    for row in cells[1:]:
        valores = [c.value for c in row]
        if all(v is None for v in valores):
            continue
        linhas.append(dict(zip(header, valores)))
    return linhas


def carregar_estrutura_precos(conn, xlsm: Path, fonte: str = "bi_matriz") -> int:
    """Carrega silver.d_estrutura (matriz de preço/estoque por unidade) das abas
    Matriz_XX de base_precos.xlsm (BI V.2/BI Matriz) — task 6.4 do roadmap. Cada aba
    é 1 empreendimento; a Excel Table tem o mesmo layout em todas (ver
    _bi_ref/M_Empreendimentos.md, TABELA d_estrutura). Dedup por Código Interno
    (ING-08, primeiro-vence, mesmo critério dos demais loaders).

    `fonte` escolhe qual das DUAS matrizes de preço que a empresa mantém (R22):
      "bi_matriz" (padrão) — `BI V.2/BI Matriz/base_precos.xlsm`, usada desde a 6.4.
      "legado"            — `Preço/Apoio/Apoio - BI de Preço.xlsm`, a que alimenta
                            o PBIX "BI Preço". Use esta pro relatório novo bater
                            com os números que a gestão já conhece.
    Diferenças de layout do legado tratadas aqui: colunas de posição chamadas
    `Prumada`/`Frente/Fundo`/`Final` (produtos verticais mais antigos), bloco em
    `Torre`, e **sem coluna `codigo_cv`** — resolvido pelo nome do produto.
    """
    import openpyxl
    from psycopg import sql

    if fonte not in ("bi_matriz", "legado"):
        raise ValueError(f"fonte inválida: {fonte!r} (use 'bi_matriz' ou 'legado')")

    wb = openpyxl.load_workbook(xlsm, data_only=True)
    origem_txt = f"SharePoint: {xlsm.name} (abas Matriz_*, fonte={fonte})"

    # Tabelas "Matriz_*" com problema conhecido de origem — achado na validação
    # do relatório "Vendas Geral" (ago/2026): a aba QBV2 tem cabeçalhos com erro
    # de digitação ("Unidde"/"ID_Prço"/"Áre Privtiv"/"Permut" em vez de
    # Unidade/ID_Preço/Área Privativa/Permuta) que zeravam a coluna Unidade pro
    # produto inteiro (Quinta da Boa Vista, VSO saía 0%). A aba QBV (tabela com
    # nome mangulado "Matriz_F162427" — artefato de copiar/colar do Excel) tem o
    # MESMO produto/codigo_cv, mesma contagem de linhas, com os cabeçalhos certos
    # — é a versão boa. Pular a quebrada evita que o dedup por Código Interno
    # (primeiro-vence, ING-08) fique com a errada por causa da ordem das abas.
    # ⚠️ No arquivo do LEGADO "Matriz_QBV" é a tabela BOA (e única) do produto —
    # ignorá-la lá derrubaria Quinta da Boa Vista inteira.
    ABAS_TABELA_IGNORAR = {"Matriz_QBV"} if fonte == "bi_matriz" else set()

    # Variação de nome de coluna entre abas do mesmo workbook (cada produto foi
    # montado por cópia manual da matriz e divergiu um pouco) — normaliza pro
    # nome canônico antes de extrair os campos.
    ALIAS_COLUNA = {
        "Unidde": "Unidade",
        "ID_Prço": "ID_Preço",
        "Áre Privtiv": "Área Privativa",
        "Permut": "Permuta",
        "Tipologi": "CONFIG_1",
        "Tipologia": "CONFIG_1",
        "CONFIG 1": "CONFIG_1",
        # layout do arquivo legado (produtos verticais mais antigos)
        "Prumada": "CONFIG_1",
        "Frente/Fundo": "CONFIG_2",
        "Final": "CONFIG_3",
        "Torre": "Bloco",
    }

    # Erros de digitação no VALOR das colunas de posição (CONFIG_1/2/3) — mesma
    # família do R19, cada aba foi montada por cópia manual. Não são cosméticos:
    # na matriz do BI de Preço cada grafia vira uma COLUNA SEPARADA, com 1-2
    # unidades cada, empurrando as colunas boas para fora da tela. Levantado em
    # 13/ago/2026 sobre as 79 combinações distintas de config_* da gold.
    ALIAS_VALOR_CONFIG = {
        "LOTE MISTO (GURIT E COMERCIL)": "LOTE MISTO (GUARITA E COMERCIAL)",  # faltam letras
        "OTE MISTO":                     "LOTE MISTO",                        # falta o L
        "Muro lateral":                  "Muro Lateral",                      # caixa divergente
    }
    # NÃO incluído de propósito: config_3 = "Lateral" (1 unidade em Villas do Pq.
    # Lotes Mistos). É provável que seja "Muro Lateral" abreviado, mas "Lateral"
    # também é um valor LEGÍTIMO de config_2 (face, em Parc das Artes) — juntar
    # sem confirmar seria inventar dado. Confirmar com o backoffice.

    def _num(v):
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None  # ex.: '#N/A' / '#DIV/0!' de fórmula quebrada na planilha

    def _txt(v):
        """Texto normalizado: colapsa espaços repetidos (a origem tem
        '154  e 155 (PCD)' com espaço duplo, que também virava coluna própria)."""
        if v is None:
            return None
        s = re.sub(r"\s+", " ", str(v)).strip()
        return s or None

    def _cfg(v):
        """_txt + de-para de erro de digitação, para as colunas de posição."""
        s = _txt(v)
        return ALIAS_VALOR_CONFIG.get(s, s)

    registros: dict[str, tuple] = {}
    n_abas = 0
    for ws in wb.worksheets:
        for nome_tabela in list(ws.tables):
            if not nome_tabela.startswith("Matriz_") or nome_tabela in ABAS_TABELA_IGNORAR:
                continue
            n_abas += 1
            for r_bruto in _tabela_openpyxl(ws, nome_tabela):
                # Renomeia pelos alias SEM sobrescrever: duas colunas da mesma aba
                # podem cair no mesmo nome canônico (a aba QBV tem "Tipologia" E
                # "CONFIG 1"; o legado tem "Prumada" em algumas e "CONFIG_1" em
                # outras). Vale o primeiro valor NÃO VAZIO — antes o último vencia
                # e, quando ele era o vazio, perdia-se a coluna inteira.
                r: dict = {}
                for k, v in r_bruto.items():
                    alvo = ALIAS_COLUNA.get(k, k)
                    if r.get(alvo) in (None, "") or alvo not in r:
                        if not (v in (None, "") and alvo in r):
                            r[alvo] = v
                cod_interno = _txt(r.get("Código Interno"))
                if not cod_interno or cod_interno in registros:
                    continue
                registros[cod_interno] = (
                    cod_interno,
                    int(r["codigo_cv"]) if r.get("codigo_cv") is not None else None,
                    _txt(r.get("Produto")),
                    _txt(r.get("Bloco")),
                    _txt(r.get("Unidade")),
                    _txt(r.get("ID_Preço")),
                    _num(r.get("Área Privativa")),
                    _cfg(r.get("CONFIG_1")),
                    _cfg(r.get("CONFIG_2")),
                    _cfg(r.get("CONFIG_3")),
                    (r.get("Permuta") == "Permuta"),
                    _num(r.get("Preço")),
                    _num(r.get("Preço M²")),
                )
    wb.close()

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.d_estrutura")
        for reg in registros.values():
            cur.execute(
                sql.SQL("INSERT INTO silver.d_estrutura "
                        "(codigo_interno, codigo_cv, produto, bloco, unidade, id_preco, "
                        "area_privativa, config_1, config_2, config_3, permuta, preco, preco_m2, _origem) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING"),
                reg + (origem_txt,),
            )
        if fonte == "legado":
            # O arquivo do legado não tem coluna codigo_cv — resolve pelo NOME do
            # produto, conformado dos dois lados pela mesma função que a gold usa
            # (absorve "FIUSA 016" x "Fiusa 016", acento, caixa).
            cur.execute("""
                UPDATE silver.d_estrutura e
                   SET codigo_cv = d.codigo_interno_empreendimento
                  FROM gold.dim_empreendimento d
                 WHERE e.codigo_cv IS NULL
                   AND d.codigo_interno_empreendimento IS NOT NULL
                   AND lower(btrim(d.empreendimento_conformado COLLATE "und-x-icu")) =
                       lower(btrim(silver.conformar_empreendimento(e.produto) COLLATE "und-x-icu"))
            """)
            cur.execute("SELECT count(*) FROM silver.d_estrutura WHERE codigo_cv IS NULL")
            sem_cv = cur.fetchone()[0]
            if sem_cv:
                cur.execute("""SELECT DISTINCT produto FROM silver.d_estrutura
                                WHERE codigo_cv IS NULL ORDER BY 1""")
                nomes = ", ".join(p for (p,) in cur.fetchall())
                log.warning("  %d unidades sem codigo_cv (produto sem match em "
                            "dim_empreendimento): %s", sem_cv, nomes)
    log.info("  silver.%-24s <- %4d linhas (%d abas Matriz_*, fonte=%s)",
             "d_estrutura", len(registros), n_abas, fonte)
    return len(registros)


def carregar_metas_empreendimentos(conn, xlsx: Path) -> int:
    """Carrega silver.d_metas_empreendimentos (metas/forecast mensais) da tabela
    meta_2 (Meta.xlsx, aba base_meta) — task 6.4. Não vem da API CVDW: planejamento
    manual da gestão (R5). Grão = codigo_cv x mês x status_meta (Start/Replan).
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["base_meta"]
    linhas = _tabela_openpyxl(ws, "meta_2")
    wb.close()
    origem_txt = f"SharePoint: {xlsx.name} (aba base_meta, tabela meta_2)"

    def _int(v):
        try:
            return int(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _num(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _data(v):
        return v.date() if hasattr(v, "date") else v

    registros = []
    vistos = set()
    for r in linhas:
        cod_cv = _int(r.get("codigo_cv"))
        data = _data(r.get("data"))
        status = r.get("status_meta")
        if cod_cv is None or data is None or not status:
            continue
        chave = (cod_cv, data, status)
        if chave in vistos:
            continue
        vistos.add(chave)
        registros.append((
            cod_cv, data, status,
            r.get("chave_gv"), _data(r.get("data_base")), _int(r.get("mes")),
            r.get("empreendimento"), r.get("status_empreendimento"), r.get("regional"),
            _num(r.get("meta_house")), _num(r.get("meta_imobiliaria")),
            _num(r.get("meta_gv_house")), _num(r.get("meta_gv_imob")),
            _int(r.get("meta_qtd")), _num(r.get("meta_vgv")),
            _int(r.get("forecast_qtd")), _num(r.get("forecast_vgv")),
            _num(r.get("meta digital rpo")), _num(r.get("meta digital regional")),
            _num(r.get("meta digital")), _int(r.get("qtd apresentação")),
            _num(r.get("vgv apresentação")),
        ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.d_metas_empreendimentos")
        for reg in registros:
            cur.execute(
                sql.SQL("INSERT INTO silver.d_metas_empreendimentos "
                        "(codigo_cv, data, status_meta, chave_gv, data_base, mes, empreendimento, "
                        "status_empreendimento, regional, meta_house, meta_imobiliaria, meta_gv_house, "
                        "meta_gv_imob, meta_qtd, meta_vgv, forecast_qtd, forecast_vgv, meta_digital_rpo, "
                        "meta_digital_regional, meta_digital, qtd_apresentacao, vgv_apresentacao, _origem) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                        "ON CONFLICT DO NOTHING"),
                reg + (origem_txt,),
            )
    log.info("  silver.%-24s <- %4d linhas (de %d na tabela meta_2)",
              "d_metas_empreendimentos", len(registros), len(linhas))
    return len(registros)


def carregar_viabilidade(conn, xlsx: Path) -> int:
    """Carrega silver.d_viabilidade (parâmetros de margem por empreendimento, EAV) da
    tabela tab_viabil_padrão (d_para empreendimentos.xlsx, aba viabil_padrão) — task
    6.4. Resolve R4: no legado eram ~12 conjuntos de constantes coladas no DAX.
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["viabil_padrão"]
    linhas = _tabela_openpyxl(ws, "tab_viabil_padrão")
    wb.close()
    origem_txt = f"SharePoint: {xlsx.name} (aba viabil_padrão, tabela tab_viabil_padrão)"

    registros = []
    vistos = set()
    for r in linhas:
        cod_cv = r.get("cod_cv")
        tipo = (r.get("Tipo") or "").strip()
        if cod_cv is None or not tipo:
            continue
        chave = (int(cod_cv), tipo)
        if chave in vistos:
            continue
        vistos.add(chave)
        registros.append((int(cod_cv), tipo, r.get("Valor"), r.get("%")))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.d_viabilidade")
        for reg in registros:
            cur.execute(
                sql.SQL("INSERT INTO silver.d_viabilidade (codigo_cv, tipo, valor, percentual, _origem) "
                        "VALUES (%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING"),
                reg + (origem_txt,),
            )
    log.info("  silver.%-24s <- %4d linhas (de %d na tabela)", "d_viabilidade", len(registros), len(linhas))
    return len(registros)


def carregar_empreendimento_legado(conn, xlsx: Path) -> int:
    """Carrega silver.d_empreendimento_legado (Data Lançamento/Tipo Produto por
    empreendimento) da tabela base_cv (mesmo d_para empreendimentos.xlsx, aba
    "d_para empreendimentos"). Fonte de Data Lançamento pro cálculo do IVV padrão
    (gold.dim_ivv_padrao) — não vem da API CVDW.
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["d_para empreendimentos"]
    linhas = _tabela_openpyxl(ws, "base_cv")
    wb.close()
    origem_txt = f"SharePoint: {xlsx.name} (aba d_para empreendimentos, tabela base_cv)"

    def _data(v):
        return v.date() if hasattr(v, "date") else v

    registros = []
    vistos = set()
    for r in linhas:
        cod_cv = r.get("codigo_cv")
        if cod_cv is None or int(cod_cv) in vistos:
            continue
        vistos.add(int(cod_cv))
        registros.append((
            int(cod_cv), r.get("EP"), r.get("Empreendimentos"), r.get("Regional"),
            _data(r.get("Data Lançamento")), r.get("Tipo Produto"), r.get("Assinatura"),
        ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.d_empreendimento_legado")
        for reg in registros:
            cur.execute(
                sql.SQL("INSERT INTO silver.d_empreendimento_legado "
                        "(codigo_cv, ep, empreendimento, regional, data_lancamento, tipo_produto, assinatura, _origem) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING"),
                reg + (origem_txt,),
            )
    log.info("  silver.%-24s <- %4d linhas (de %d na tabela base_cv)",
              "d_empreendimento_legado", len(registros), len(linhas))
    return len(registros)


def carregar_ivv_padrao(conn, xlsx: Path) -> int:
    """Carrega silver.d_ivv (curva padrão de IVV acumulado por mês desde o
    lançamento) da tabela base_cv4 (mesmo d_para empreendimentos.xlsx, aba
    "IVV_padrão"). Formato largo no Excel (colunas "1".."36" = mês desde o
    lançamento); despivotado aqui pro grão codigo_cv x mês.
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["IVV_padrão"]
    linhas = _tabela_openpyxl(ws, "base_cv4")
    wb.close()
    origem_txt = f"SharePoint: {xlsx.name} (aba IVV_padrão, tabela base_cv4)"

    registros = []
    for r in linhas:
        cod_cv = r.get("codigo_cv")
        if cod_cv is None:
            continue
        for mes in range(1, 37):
            valor = r.get(str(mes))
            if valor is None:
                continue
            registros.append((
                int(cod_cv), r.get("EP"), r.get("Empreendimentos"), r.get("Regional"),
                r.get("Assinatura"), mes, float(valor),
            ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.d_ivv")
        for reg in registros:
            cur.execute(
                sql.SQL("INSERT INTO silver.d_ivv "
                        "(codigo_cv, ep, empreendimento, regional, assinatura, mes, pct_ivv, _origem) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING"),
                reg + (origem_txt,),
            )
    log.info("  silver.%-24s <- %4d linhas (de %d empreendimentos x 36 meses)",
              "d_ivv", len(registros), len(linhas))
    return len(registros)


def carregar_distratos_2025(conn, xlsx: Path) -> int:
    """Carrega silver.distratos_2025 (detalhe financeiro de distrato: multa, valor
    pago, devolução, parcelas — não existe na API CVDW) da aba "Base Distratos" de
    relatorio_distratos.xlsx. Sheet cru (sem Excel Table nomeada), header na linha 1
    com quebra de linha em alguns nomes de coluna — normaliza antes de casar.
    """
    import openpyxl
    from psycopg import sql

    def _cab(v) -> str:
        return " ".join(str(v or "").split())

    def _num(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _int(v):
        try:
            return int(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _data(v):
        return v.date() if hasattr(v, "date") else v

    def _txt(v):
        if v is None:
            return None
        s = str(v).strip()
        return s or None

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Base Distratos"]
    linhas = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    origem_txt = f"SharePoint: {xlsx.name} (aba Base Distratos)"

    hdr = {_cab(v): i for i, v in enumerate(linhas[0]) if v is not None}

    def g(r, nome):
        i = hdr.get(nome)
        return r[i] if i is not None and i < len(r) else None

    registros = []
    for r in linhas[1:]:
        if all(v is None for v in r):
            continue
        registros.append((
            _int(g(r, "Contrato")),
            _txt(g(r, "Filial")),
            _txt(g(r, "Bloco")),
            _txt(g(r, "Unidade")),
            _txt(g(r, "Cliente")),
            _data(g(r, "Data do Contrato")),
            _data(g(r, "Data do Distrato")),
            _num(g(r, "Valor de Venda")),
            _num(g(r, "Valor do Contrato")),
            _num(g(r, "Área Privativa")),
            _num(g(r, "Valor Pago")),
            _num(g(r, "Valor Pago Atualizado")),
            _num(g(r, "Valor Multa")),
            _int(g(r, "Fruição")),
            _num(g(r, "Valor de Devolução")),
            _txt(g(r, "Forma de Devolução")),
            _int(g(r, "Nº de Parc.")),
            _txt(g(r, "Status Contrato")),
            _txt(g(r, "Tipo Cto.")),
            _txt(g(r, "Trans. Fil.")),
            _txt(g(r, "Motivo 1")),
            _txt(g(r, "Motivo 2")),
            _txt(g(r, "Gerente Responsavel")),
        ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.distratos_2025")
        for reg in registros:
            cur.execute(
                sql.SQL("INSERT INTO silver.distratos_2025 "
                        "(contrato, produto, bloco, unidade, cliente, data_contrato, data_distrato, "
                        "valor_venda, valor_contrato, area_privativa, valor_pago, valor_pago_atualizado, "
                        "valor_multa, fruicao, valor_devolucao, forma_devolucao, numero_parcelas, "
                        "status_contrato, tipo_contrato, trans_fil, motivo_1, motivo_2, "
                        "gerente_responsavel, _origem) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"),
                reg + (origem_txt,),
            )
    log.info("  silver.%-24s <- %4d linhas (de %d na aba Base Distratos)",
              "distratos_2025", len(registros), len(linhas) - 1)
    return len(registros)


def carregar_depara_produtos(conn, xlsm: Path) -> int:
    """Carrega o de-para de empreendimento da aba DE_PARA_PRODUTOS (Vendas Consolidadas.xlsm).

    Colunas B/C/D: Produto (nome CRM) -> Produto_Depara (conformado) -> EP (espaço de negócios).
    """
    import openpyxl  # dependência opcional, só p/ ler o .xlsm
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    ws = wb["DE_PARA_PRODUTOS"]
    linhas = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(linhas) if r and "Produto" in r and "Produto_Depara" in r)
    hdr = linhas[hdr_idx]
    jo, jc, je = hdr.index("Produto"), hdr.index("Produto_Depara"), hdr.index("EP")

    registros = []
    vistos = set()
    for r in linhas[hdr_idx + 1:]:
        origem = (r[jo] or "").strip() if r[jo] else ""
        if not origem or origem in vistos:
            continue
        vistos.add(origem)
        registros.append((origem,
                          (r[jc] or "").strip() if r[jc] else None,
                          (r[je] or "").strip() if r[je] else None))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.dpara_empreendimento")
        for origem, conf, ep in registros:
            cur.execute(
                sql.SQL("INSERT INTO silver.dpara_empreendimento (nome_origem, nome_conformado, ep) "
                        "VALUES (%s, %s, %s) ON CONFLICT (nome_origem) DO NOTHING"),
                (origem, conf, ep),
            )
    log.info("  silver.%-22s <- %3d linhas (aba DE_PARA_PRODUTOS)", "dpara_empreendimento", len(registros))
    return len(registros)


def carregar_gerentes(conn, xlsx: Path) -> int:
    """Carrega os 2 de-paras de classificação do depara_gerentes.xlsx (fonte autoritativa).

    Aba "contexto"    -> silver.dpara_gerente_contexto   (classifica a RESERVA, pela
        chave crua "Gerente Responsavel" do CVCRM): Gerente Responsável | Gerente
        Apelido | Share | House/Parcerias | Regional.
    Aba "imobiliaria" -> silver.dpara_imobiliaria_house  (classifica LEADS e
        PRÉ-CADASTROS, pelo escritório do corretor vindo do headcount):
        Imobiliária | Share | House/Parcerias | Regional.

    Nenhuma das duas é fonte de EQUIPE do corretor — isso vem do headcount
    (ver carregar_headcount_corretores).
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    ws_c = wb["contexto"]
    linhas_c = list(ws_c.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(linhas_c)
                   if r and any(c == "Gerente Responsável" for c in r if c))
    hdr = list(linhas_c[hdr_idx])
    j_resp, j_apc = hdr.index("Gerente Responsável"), hdr.index("Gerente Apelido")
    j_share, j_hp, j_reg = hdr.index("Share"), hdr.index("House/Parcerias"), hdr.index("Regional")

    contextos, vistos_c = [], set()
    for r in linhas_c[hdr_idx + 1:]:
        resp = (r[j_resp] or "").strip() if r[j_resp] else ""
        if not resp or resp in vistos_c:
            continue
        vistos_c.add(resp)
        contextos.append((
            resp,
            (r[j_apc] or "").strip() if r[j_apc] else None,
            (r[j_share] or "").strip() if r[j_share] else None,
            (r[j_hp] or "").strip() if r[j_hp] else None,
            (r[j_reg] or "").strip() if r[j_reg] else None,
        ))

    # --- aba "imobiliaria": escritório -> Share/House/Regional (leads + pré-cadastros)
    ws_i = wb["imobiliaria"]
    linhas_i = list(ws_i.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(linhas_i)
                   if r and any(c == "Imobiliária" for c in r if c))
    hdr = list(linhas_i[hdr_idx])
    j_imob, j_sh = hdr.index("Imobiliária"), hdr.index("Share")
    j_hp, j_reg = hdr.index("House/Parcerias"), hdr.index("Regional")

    imobs, vistos_i = [], set()
    for r in linhas_i[hdr_idx + 1:]:
        imob = (r[j_imob] or "").strip() if r[j_imob] else ""
        if not imob or imob.lower() in vistos_i:
            continue
        vistos_i.add(imob.lower())
        imobs.append((
            imob,
            (r[j_sh] or "").strip() if r[j_sh] else None,
            (r[j_hp] or "").strip() if r[j_hp] else None,
            (r[j_reg] or "").strip() if r[j_reg] else None,
        ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.dpara_gerente_contexto")
        for resp, ap_, sh, hp, reg in contextos:
            cur.execute(
                sql.SQL("INSERT INTO silver.dpara_gerente_contexto "
                        "(gerente_responsavel, gerente_apelido, share, house_parcerias, regional) "
                        "VALUES (%s,%s,%s,%s,%s) ON CONFLICT (gerente_responsavel) DO NOTHING"),
                (resp, ap_, sh, hp, reg),
            )
        cur.execute("TRUNCATE silver.dpara_imobiliaria_house")
        for imob, sh, hp, reg in imobs:
            cur.execute(
                sql.SQL("INSERT INTO silver.dpara_imobiliaria_house "
                        "(imobiliaria, share, house_parcerias, regional) "
                        "VALUES (%s,%s,%s,%s) ON CONFLICT (imobiliaria) DO NOTHING"),
                (imob, sh, hp, reg),
            )
    log.info("  silver.%-24s <- %3d linhas (aba contexto, autoritativo)", "dpara_gerente_contexto", len(contextos))
    log.info("  silver.%-24s <- %3d linhas (aba imobiliaria, autoritativo)", "dpara_imobiliaria_house", len(imobs))
    return len(contextos) + len(imobs)


def carregar_headcount_corretores(conn, xlsx: Path) -> int:
    """Carrega silver.dpara_corretor_headcount da planilha manual do backoffice
    (Base Corretores Pafil, aba "Base Pafil") — fonte autoritativa de Gerente/
    Supervisor por corretor. Só carrega os corretores com "Ativo/Inativo" = Ativo
    (o pedido do dev: "usar o nome do corretor como chave, referenciando corretores
    ativos da tabela de headcount" — o resto da planilha é histórico de desligados).
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Base Pafil"]
    linhas = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(linhas) if r and any(c == "Nome" for c in r if c) and any(c == "Gerente" for c in r if c))
    hdr = list(linhas[hdr_idx])
    j_house, j_nome, j_apelido = hdr.index("House"), hdr.index("Nome"), hdr.index("Apelido")
    j_imob, j_ger, j_sup = hdr.index("Imobiliaria"), hdr.index("Gerente"), hdr.index("Supervisor")
    j_ativo = hdr.index("Ativo/Inativo")

    regs, vistos = [], set()
    for r in linhas[hdr_idx + 1:]:
        nome = (r[j_nome] or "").strip() if r[j_nome] else ""
        ativo = (r[j_ativo] or "").strip().lower() if r[j_ativo] else ""
        if not nome or ativo != "ativo" or nome.lower() in vistos:
            continue
        vistos.add(nome.lower())
        regs.append((
            nome,
            (r[j_apelido] or "").strip() if r[j_apelido] else None,
            (r[j_imob] or "").strip() if r[j_imob] else None,
            (r[j_house] or "").strip() if r[j_house] else None,
            (r[j_ger] or "").strip() if r[j_ger] else None,
            (r[j_sup] or "").strip() if r[j_sup] else None,
        ))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.dpara_corretor_headcount")
        for nome, apelido, imob, house, ger, sup in regs:
            cur.execute(
                sql.SQL("INSERT INTO silver.dpara_corretor_headcount "
                        "(corretor, apelido, imobiliaria, house, gerente, supervisor) "
                        "VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (corretor) DO NOTHING"),
                (nome, apelido, imob, house, ger, sup),
            )
    log.info("  silver.%-24s <- %3d linhas (aba Base Pafil, só Ativos)", "dpara_corretor_headcount", len(regs))
    return len(regs)


def _bloco_apoio(linhas: list, hdr_idx: int, ancora: str, offsets: dict[str, int]) -> list[dict]:
    """Extrai um "bloco" de de-para da aba Apoio (vários de-paras lado a lado).

    ancora  = nome de coluna ÚNICO no header que identifica o bloco;
    offsets = {nome_campo: deslocamento relativo à âncora} (pode ser negativo).
    """
    hdr = list(linhas[hdr_idx])
    j0 = hdr.index(ancora)
    regs = []
    for r in linhas[hdr_idx + 1:]:
        reg = {}
        for campo, off in offsets.items():
            v = r[j0 + off] if j0 + off < len(r) else None
            reg[campo] = str(v).strip() if v is not None and str(v).strip() != "" else None
        regs.append(reg)
    return regs


def carregar_leads_apoio(conn, xlsm: Path, secoes: dict[str, str]) -> int:
    """Recarrega os 4 de-paras de lead da aba Apoio do Base de Leads.xlsm (fonte VIVA).

    A aba Apoio é o painel que o analista já mantém hoje; cada de-para é um bloco de
    colunas lado a lado (header na mesma linha):
      - Situação/Qualificado?            -> silver.dpara_qualificacao_lead  (MQL)
      - Canal_Mídia/Canal/Mídia          -> silver.dpara_canal_midia        (origem+mídia)
      - Canal/Lead ou Prospect           -> silver.dpara_ativo_receptivo
      - CONCAT/CANAL/MÍDIA/ORIGEM DA...  -> silver.dpara_canal_midia_dc     (chave UTM, canal 2.0)
    Complemento canal/mídia: concats do JSON out/24 ausentes no xlsm entram marcados
    (leads antigos perderiam o match sem a união).
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    ws = wb["Apoio"]
    linhas = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(linhas) if r and any(c == "Canal_Mídia" for c in r if c))
    origem_txt = f"SharePoint: {xlsm.name} (aba Apoio)"
    total = 0

    with conn.cursor() as cur:
        # --- qualificação (MQL): Situação -> Qualificado? --------------------
        regs = [r for r in _bloco_apoio(linhas, hdr_idx, "Qualificado?", {"situacao": -1, "qualificado": 0})
                if r["situacao"]]
        vistos: set[str] = set()
        cur.execute("TRUNCATE silver.dpara_qualificacao_lead")
        for r in regs:
            if r["situacao"].lower() in vistos:
                continue
            vistos.add(r["situacao"].lower())
            cur.execute("INSERT INTO silver.dpara_qualificacao_lead (situacao, qualificado, _origem) "
                        "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                        (r["situacao"], r["qualificado"], origem_txt))
        log.info("  silver.%-24s <- %3d linhas (aba Apoio, vivo)", "dpara_qualificacao_lead", len(vistos))
        total += len(vistos)

        # --- ativo/receptivo: Canal -> Lead ou Prospect ----------------------
        regs = [r for r in _bloco_apoio(linhas, hdr_idx, "Lead ou Prospect", {"canal": -1, "lead_ou_prospect": 0})
                if r["canal"]]
        vistos = set()
        cur.execute("TRUNCATE silver.dpara_ativo_receptivo")
        for r in regs:
            if r["canal"].lower() in vistos:
                continue
            vistos.add(r["canal"].lower())
            cur.execute("INSERT INTO silver.dpara_ativo_receptivo (canal, lead_ou_prospect, _origem) "
                        "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                        (r["canal"], r["lead_ou_prospect"], origem_txt))
        log.info("  silver.%-24s <- %3d linhas (aba Apoio, vivo)", "dpara_ativo_receptivo", len(vistos))
        total += len(vistos)

        # --- canal/mídia por origem+mídia (vocabulário out24, vivo) ----------
        regs = [r for r in _bloco_apoio(linhas, hdr_idx, "Canal_Mídia", {"concat": 0, "canal": 1, "midia": 3})
                if r["concat"]]
        vistos = set()
        cur.execute("TRUNCATE silver.dpara_canal_midia")
        for r in regs:
            if r["concat"].lower() in vistos:
                continue
            vistos.add(r["concat"].lower())
            cur.execute("INSERT INTO silver.dpara_canal_midia (concat, canal, midia, vigencia_de, _origem) "
                        "VALUES (%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                        (r["concat"], r["canal"], r["midia"], "2024-10-01", origem_txt))
        n_xlsm = len(vistos)
        n_json = 0
        if "dpara_canal_midia_out24" in secoes:  # complemento: foto out/24 do M legado
            for concat, canal, midia in _decodificar(secoes["dpara_canal_midia_out24"]):
                concat = (concat or "").strip()
                if not concat or concat.lower() in vistos:
                    continue
                vistos.add(concat.lower())
                cur.execute("INSERT INTO silver.dpara_canal_midia (concat, canal, midia, vigencia_de, _origem) "
                            "VALUES (%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                            (concat, canal, midia, "2024-10-01",
                             "JSON out24 (complemento — AUSENTE no xlsm; migrar p/ o Excel)"))
                n_json += 1
        log.info("  silver.%-24s <- %3d linhas (%d aba Apoio + %d herdadas do JSON out24)",
                 "dpara_canal_midia", n_xlsm + n_json, n_xlsm, n_json)
        total += n_xlsm + n_json

        # --- canal/mídia D.C (chave UTM — decide o canal 2.0 pós-out/24) -----
        regs = [r for r in _bloco_apoio(linhas, hdr_idx, "CONCAT",
                                        {"concat": 0, "canal": 1, "midia": 2, "trafego": 3})
                if r["concat"]]
        vistos = set()
        cur.execute("TRUNCATE silver.dpara_canal_midia_dc")
        for r in regs:
            if r["concat"].lower() in vistos:
                continue
            vistos.add(r["concat"].lower())
            cur.execute("INSERT INTO silver.dpara_canal_midia_dc (concat, canal, midia, trafego, _origem) "
                        "VALUES (%s,%s,%s,%s,%s) ON CONFLICT DO NOTHING",
                        (r["concat"], r["canal"], r["midia"], r["trafego"], origem_txt))
        log.info("  silver.%-24s <- %3d linhas (aba Apoio, vivo)", "dpara_canal_midia_dc", len(vistos))
        total += len(vistos)

    return total


def carregar_etapa_precadastro(conn, xlsm: Path) -> int:
    """Carrega silver.dpara_etapa_precadastro (DP-05) da aba Apoio do Base - Crédito.xlsm.

    Mesmo padrão de bloco lado-a-lado do Base de Leads.xlsm: colunas
    "Etapa WKF" | "Etapa precadastro BI" | "Etapa precadastro BI Detalhada"
    (a situação crua do CVCRM -> etapa do funil de crédito do BI).
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    ws = wb["Apoio"]
    linhas = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(linhas) if r and any(c == "Etapa precadastro BI" for c in r if c))
    origem_txt = f"SharePoint: {xlsm.name} (aba Apoio)"

    regs = [r for r in _bloco_apoio(
                linhas, hdr_idx, "Etapa precadastro BI",
                {"etapa_wkf": -1, "etapa_bi": 0, "etapa_bi_detalhada": 1})
            if r["etapa_wkf"]]

    vistos: set[str] = set()
    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.dpara_etapa_precadastro")
        for r in regs:
            chave = r["etapa_wkf"].lower()
            if chave in vistos:
                continue
            vistos.add(chave)
            cur.execute(
                sql.SQL("INSERT INTO silver.dpara_etapa_precadastro "
                        "(etapa_wkf, etapa_bi, etapa_bi_detalhada, _origem) "
                        "VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING"),
                (r["etapa_wkf"], r["etapa_bi"], r["etapa_bi_detalhada"], origem_txt),
            )
    log.info("  silver.%-24s <- %3d linhas (aba Apoio)", "dpara_etapa_precadastro", len(vistos))
    return len(vistos)


def carregar_precadastro_credito_manual(conn, xlsx: Path) -> int:
    """Carrega silver.precadastros_credito_manual do export "Relatório Web" do
    CVCRM (relatorios_precadastro.xlsx, aba Sheet1). NÃO é de-para — 1 linha
    por pré-cadastro (chave = Id), só as 2 colunas que a API CVDW não traz:
    "Aprovação de crédito" e "Encaminhado ao CCA" (a maioria fica NULL, é
    esperado — cobre só os pré-cadastros tocados pelo time de crédito).
    """
    import openpyxl
    from psycopg import sql

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    linhas = ws.iter_rows(values_only=True)
    hdr = list(next(linhas))
    j_id = hdr.index("Id")
    j_apr = hdr.index("Aprovação de crédito")
    j_cca = hdr.index("Encaminhado ao CCA")
    origem_txt = f"SharePoint: {xlsx.name} (aba Sheet1)"

    regs = []
    vistos: set[int] = set()
    for r in linhas:
        id_pc = r[j_id]
        if id_pc is None or id_pc in vistos:
            continue
        vistos.add(id_pc)
        apr = (r[j_apr] or "").strip() if r[j_apr] else None
        cca = (r[j_cca] or "").strip() if r[j_cca] else None
        regs.append((int(id_pc), apr, cca))

    with conn.cursor() as cur:
        cur.execute("TRUNCATE silver.precadastros_credito_manual")
        for id_pc, apr, cca in regs:
            cur.execute(
                sql.SQL("INSERT INTO silver.precadastros_credito_manual "
                        "(id_precadastro, aprovacao_credito, encaminhado_cca, _origem) "
                        "VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING"),
                (id_pc, apr, cca, origem_txt),
            )
    log.info("  silver.%-24s <- %3d linhas (aba Sheet1)", "precadastros_credito_manual", len(regs))
    return len(regs)


def main() -> int:
    ap = argparse.ArgumentParser(description="Popula seeds de-para (JSON legado + opcional xlsm).")
    ap.add_argument("--xlsm", help="caminho de Vendas Consolidadas.xlsm (carrega DE_PARA_PRODUTOS)")
    ap.add_argument("--gerentes", help="caminho de depara_gerentes.xlsx (recarrega "
                                       "dpara_gerente_contexto). Default: variável DEPARA_GERENTES_XLSX do .env")
    ap.add_argument("--headcount-corretores", help="caminho de Base Corretores Pafil.xlsx (recarrega "
                                       "dpara_corretor_headcount, só Ativos). Default: HEADCOUNT_CORRETORES_XLSX do .env")
    ap.add_argument("--leads-apoio", help="caminho de Base de Leads.xlsm (aba Apoio: recarrega os 4 "
                                          "de-paras de lead vivos). Default: DEPARA_LEADS_XLSM do .env")
    ap.add_argument("--etapa-precadastro", help="caminho de Base - Crédito.xlsm (aba Apoio: recarrega "
                                          "dpara_etapa_precadastro, DP-05). Default: ETAPA_PRECADASTRO_XLSM do .env")
    ap.add_argument("--credito-manual", help="caminho de relatorios_precadastro.xlsx (recarrega "
                                          "precadastros_credito_manual: Aprovação de crédito/Encaminhado ao CCA). "
                                          "Default: PRECADASTRO_CREDITO_XLSX do .env")
    ap.add_argument("--estrutura-precos", help="caminho da matriz de preço/estoque por unidade "
                                          "(recarrega silver.d_estrutura — task 6.4). "
                                          "Default: ESTRUTURA_PRECOS_XLSM do .env")
    ap.add_argument("--estrutura-fonte", choices=("bi_matriz", "legado"), default=None,
                    help="qual das duas matrizes de preço da empresa usar (R22): 'bi_matriz' = "
                         "base_precos.xlsm (BI V.2); 'legado' = Apoio - BI de Preço.xlsm, a que "
                         "alimenta o PBIX BI Preço. Default: ESTRUTURA_PRECOS_FONTE do .env, "
                         "ou 'legado' quando o arquivo se chama 'Apoio - BI de Preço.xlsm'.")
    ap.add_argument("--metas-empreendimentos", help="caminho de Meta.xlsx (recarrega "
                                          "silver.d_metas_empreendimentos — task 6.4). "
                                          "Default: METAS_EMPREENDIMENTOS_XLSX do .env")
    ap.add_argument("--viabilidade", help="caminho de d_para empreendimentos.xlsx (recarrega "
                                          "silver.d_viabilidade, parâmetros de margem — task 6.4). "
                                          "Default: VIABILIDADE_XLSX do .env")
    ap.add_argument("--distratos-2025", help="caminho de relatorio_distratos.xlsx (recarrega "
                                          "silver.distratos_2025, detalhe financeiro de distrato). "
                                          "Default: DISTRATOS_2025_XLSX do .env")
    args = ap.parse_args()

    # Fonte autoritativa dos gerentes: arg explícito ou o xlsx sincronizado (via .env).
    gerentes_path = args.gerentes or os.getenv("DEPARA_GERENTES_XLSX")
    headcount_path = args.headcount_corretores or os.getenv("HEADCOUNT_CORRETORES_XLSX")
    leads_apoio_path = args.leads_apoio or os.getenv("DEPARA_LEADS_XLSM")
    etapa_precadastro_path = args.etapa_precadastro or os.getenv("ETAPA_PRECADASTRO_XLSM")
    credito_manual_path = args.credito_manual or os.getenv("PRECADASTRO_CREDITO_XLSX")
    estrutura_precos_path = args.estrutura_precos or os.getenv("ESTRUTURA_PRECOS_XLSM")
    # fonte da matriz de preço (R22): explícita > .env > deduzida do nome do arquivo
    estrutura_fonte = (args.estrutura_fonte or os.getenv("ESTRUTURA_PRECOS_FONTE") or "").strip()
    if not estrutura_fonte:
        estrutura_fonte = ("legado"
                           if estrutura_precos_path and "apoio" in Path(estrutura_precos_path).name.lower()
                           else "bi_matriz")
    metas_empreendimentos_path = args.metas_empreendimentos or os.getenv("METAS_EMPREENDIMENTOS_XLSX")
    viabilidade_path = args.viabilidade or os.getenv("VIABILIDADE_XLSX")
    distratos_2025_path = args.distratos_2025 or os.getenv("DISTRATOS_2025_XLSX")

    configurar_logging(False)
    if not MD_LEGADO.exists():
        log.error("Markdown legado não encontrado: %s", MD_LEGADO)
        return 2

    md = MD_LEGADO.read_text(encoding="utf-8")
    carregar.secoes = _secoes(md)  # type: ignore[attr-defined]

    cfg = carregar_config_pg()
    total = 0
    n_seeds = 0
    with db.conectar(cfg) as conn:
        # de-paras de lead vivos (aba Apoio): se o xlsm existe, ele é a fonte — não
        # sobrescrever com as fotos congeladas do JSON legado.
        SEEDS_APOIO = {"dpara_qualificacao_lead", "dpara_ativo_receptivo", "dpara_canal_midia"}
        apoio_ok = bool(leads_apoio_path) and Path(leads_apoio_path).exists()

        log.info("Populando seeds de-para a partir de %s:", MD_LEGADO.name)
        for mapa in MAPEAMENTOS:
            # havendo xlsm autoritativo, nunca sobrescreve com o JSON legado congelado
            if apoio_ok and mapa["tabela"] in SEEDS_APOIO:
                continue
            total += carregar(conn, mapa)
            n_seeds += 1
        if gerentes_path:
            gx = Path(gerentes_path)
            if gx.exists():
                total += carregar_gerentes(conn, gx)
                n_seeds += 1
            else:
                # não recarrega do JSON: mantém a dpara_gerentes atual em vez de sujá-la
                log.warning("  gerentes xlsx não encontrado (%s) — dpara_gerentes mantida como está.", gx)
        if headcount_path:
            hx = Path(headcount_path)
            if hx.exists():
                total += carregar_headcount_corretores(conn, hx)
                n_seeds += 1
            else:
                log.warning("  headcount xlsx não encontrado (%s) — dpara_corretor_headcount mantida como está.", hx)
        if apoio_ok:
            total += carregar_leads_apoio(conn, Path(leads_apoio_path), carregar.secoes)  # type: ignore[attr-defined]
            n_seeds += 4
        elif leads_apoio_path:
            log.warning("  Base de Leads.xlsm não encontrado (%s) — de-paras de lead mantidos como estão.",
                        leads_apoio_path)
        if args.xlsm:
            xlsm = Path(args.xlsm)
            if xlsm.exists():
                total += carregar_depara_produtos(conn, xlsm)
                n_seeds += 1
            else:
                log.warning("  --xlsm não encontrado: %s", xlsm)
        if etapa_precadastro_path:
            ep = Path(etapa_precadastro_path)
            if ep.exists():
                total += carregar_etapa_precadastro(conn, ep)
                n_seeds += 1
            else:
                log.warning("  etapa-precadastro xlsm não encontrado (%s) — dpara_etapa_precadastro mantida como está.", ep)
        if credito_manual_path:
            cm = Path(credito_manual_path)
            if cm.exists():
                total += carregar_precadastro_credito_manual(conn, cm)
                n_seeds += 1
            else:
                log.warning("  credito-manual xlsx não encontrado (%s) — precadastros_credito_manual mantida como está.", cm)
        if estrutura_precos_path:
            ep2 = Path(estrutura_precos_path)
            if ep2.exists():
                total += carregar_estrutura_precos(conn, ep2, fonte=estrutura_fonte)
                n_seeds += 1
            else:
                log.warning("  estrutura-precos xlsm não encontrado (%s) — d_estrutura mantida como está.", ep2)
        if metas_empreendimentos_path:
            me = Path(metas_empreendimentos_path)
            if me.exists():
                total += carregar_metas_empreendimentos(conn, me)
                n_seeds += 1
            else:
                log.warning("  metas-empreendimentos xlsx não encontrado (%s) — d_metas_empreendimentos mantida como está.", me)
        if viabilidade_path:
            vb = Path(viabilidade_path)
            if vb.exists():
                # mesmo arquivo alimenta 3 seeds: viabilidade, launch date (IVV) e a
                # curva padrão de IVV em si (abas diferentes do mesmo workbook).
                total += carregar_viabilidade(conn, vb)
                total += carregar_empreendimento_legado(conn, vb)
                total += carregar_ivv_padrao(conn, vb)
                n_seeds += 3
            else:
                log.warning("  viabilidade xlsx não encontrado (%s) — d_viabilidade/d_empreendimento_legado/d_ivv mantidas como estão.", vb)
        if distratos_2025_path:
            d25 = Path(distratos_2025_path)
            if d25.exists():
                total += carregar_distratos_2025(conn, d25)
                n_seeds += 1
            else:
                log.warning("  distratos-2025 xlsx não encontrado (%s) — distratos_2025 mantida como está.", d25)
        conn.commit()
    log.info("Concluído: %d linhas carregadas em %d seeds.", total, n_seeds)
    log.info("Pendentes (SharePoint): feriados, profissões, equipe_corretor (superada por dim_corretor).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
